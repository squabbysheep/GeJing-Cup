#coding=gbk

import pandas as pd
import jieba
import re
from collections import Counter
from wordcloud import WordCloud
from matplotlib import pyplot as plt
import numpy as np
from PIL import Image
import openpyxl
import os
import time
import random
import re

plt.rcParams['font.sans-serif'] = ['SimHei']
plt.rcParams['axes.unicode_minus'] = False

data_path = "data.xlsx"
df = pd.read_excel(data_path)

wb = openpyxl.load_workbook(data_path)
ws = wb['Data']


# ��Ƶͳ��
all_words = []

for word in df['words']:
    word_list = str(word).split(" ")
    all_words.extend(word_list)
# print(all_words)

# ����ͼ
# img_path = r'D:\Pycharm\Project\Ploting\img'
#
# img_save = os.path.join(img_path, '1.png')
#
# text = ' '.join(all_words)
#
# test_mask = np.array(Image.open(r'D:\Pycharm\Project\Ploting\picture\xin.png'))
# wordcloud = WordCloud(font_path=r'C:\Windows\Fonts\simkai.ttf', collocations=False, width=1200, height=980,
#                       mask=test_mask, scale=15).generate(text)
#
# plt.imshow(wordcloud, interpolation="bilinear")
# plt.axis("off")
#
# plt.savefig(img_save)
# plt.show()


# top20��Ƶͳ��
# x = []
# y = []
# mycount = Counter(all_words)
# for key, val in mycount.most_common(20):  # ���򣨷���ǰ100����
#     x.append(key)
#     y.append(val)
#
# fig, ax = plt.subplots()
# ax.barh(x, y, color="deepskyblue")
# labels = ax.get_xticklabels()
# plt.setp(labels, rotation=0, horizontalalignment='right')
#
# for a, b in zip(x, y):
#     plt.text(b+1, a, b, ha='center', va='center')
# ax.legend(["label"], loc="lower right")
#
# plt.rcParams['font.sans-serif'] = ['SimHei']  # ����������ʾ���ı�ǩ
# plt.ylabel('����')
# plt.xlabel('��Ƶ')
# plt.rcParams['savefig.dpi'] = 300  # ͼƬ����
# plt.rcParams['figure.dpi'] = 300  # �ֱ���
# plt.rcParams['figure.figsize'] = (15.0, 8.0)  # �ߴ�
# plt.title("Top20��Ƶͳ��")
# plt.show()
# top20��Ƶͳ��

# paper_type = df['paper_type'].value_counts()
# print(paper_type)
# plt.grid(True)
# paper_type.plot(kind='bar')
# paper_type.plot.pie(subplots=True, autopct='%4.2f%%')
# plt.show()

# ��ݴ���
# for download in df['download']:
#     m = re.findall("\d+", download)
#     print(m)
# ws["L0"] = 'year'
# for k, date in enumerate(df['date']):
#     # print(date[:4])
#     ws['L' + str(k+2)] = str(date[:4])
#
# wb.save(data_path)
# ��ݴ���

# �����������ϵͼ
# years = df['Year'].value_counts(sort=False)
#
# year_dict = dict(years)
# print(year_dict)
#
# plot_year = []
# plot_nums = []
#
# for key, value in year_dict.items():
#     plot_year.append(key)
#
# print(plot_year)
# sort_year = sorted(plot_year)
# print(sort_year)
# for year in sort_year:
#     plot_nums.append(year_dict[year])
# print(plot_nums)

# def autolabel(rects):
#     for rect in rects:
#         height = rect.get_height()
#         plt.text(rect.get_x()+rect.get_width()/2.- 0.2, 1.03*height, '%s' % int(height))
#
# autolabel(plt.bar(range(len(plot_nums)), plot_nums, color='rgb', tick_label=sort_year))
# plt.ylabel('������')
# plt.xlabel('���')
# plt.title("�����������ϵͼ")
# plt.show()

# fig, ax = plt.subplots()
# ax.barh(sort_year, plot_nums, color="deepskyblue")
# labels = ax.get_xticklabels()
# plt.setp(labels, rotation=0, horizontalalignment='right')
#
# for a, b in zip(sort_year, plot_nums):
#     plt.text(b+1, a, b, ha='center', va='center')
# ax.legend(["label"], loc="lower right")
#
# plt.rcParams['font.sans-serif'] = ['SimHei']  # ����������ʾ���ı�ǩ
# plt.ylabel('����')
# plt.xlabel('��Ƶ')
# plt.rcParams['savefig.dpi'] = 300  # ͼƬ����
# plt.rcParams['figure.dpi'] = 300  # �ֱ���
# plt.rcParams['figure.figsize'] = (15.0, 8.0)  # �ߴ�
# plt.title("���")
# plt.show()
# �����������ϵͼ

# ����������
# for j, download in enumerate(df['download']):
#     m = re.findall("\d+", download)
#     ws['M' + str(j+2)] = m[0]
#
# wb.save(data_path)
# ����������

df_num = df[df['download_num'] > 5000]
# print(df_num.head())

num_word = []
for word in df_num['words']:
    word_list = str(word).split(" ")
    num_word.extend(word_list)

# ����ͼ
# img_path = r'D:\Pycharm\Project\Ploting\img'
#
# # img_save = os.path.join(img_path, '1.png')
#
# text = ' '.join(num_word)
#
# test_mask = np.array(Image.open(r'D:\Pycharm\Project\Ploting\picture\chinamap.png'))
# wordcloud = WordCloud(font_path=r'C:\Windows\Fonts\simkai.ttf', collocations=False, width=1200, height=980,
#                       mask=test_mask, scale=15).generate(text)
#
# plt.imshow(wordcloud, interpolation="bilinear")
# plt.axis("off")
#
# # plt.savefig(img_save)
# plt.show()

# top20��Ƶͳ��
# x = []
# y = []
# mycount = Counter(num_word)
# for key, val in mycount.most_common(20):  # ���򣨷���ǰ100����
#     x.append(key)
#     y.append(val)
#
# fig, ax = plt.subplots()
# ax.barh(x, y, color="gray")
# labels = ax.get_xticklabels()
# plt.setp(labels, rotation=0, horizontalalignment='right')
#
# for a, b in zip(x, y):
#     plt.text(b+1, a, b, ha='center', va='center')
# ax.legend(["label"], loc="lower right")
#
# plt.rcParams['font.sans-serif'] = ['SimHei']  # ����������ʾ���ı�ǩ
# plt.ylabel('����')
# plt.xlabel('��Ƶ')
# plt.rcParams['savefig.dpi'] = 300  # ͼƬ����
# plt.rcParams['figure.dpi'] = 300  # �ֱ���
# plt.rcParams['figure.figsize'] = (15.0, 8.0)  # �ߴ�
# plt.title("�������ش���5000��Ƶͳ��")
# plt.show()
# top20��Ƶͳ��

# ����������
# for j, download in enumerate(df['quote']):
#     m = re.findall("\d+", download)
#     ws['N' + str(j+2)] = m[0]
#
# wb.save(data_path)
# ����������
df_quote = df[df['quote_num'] >= 100]
# print(df_num.head())

quote_word = []
for word in df_quote['words']:
    word_list = str(word).split(" ")
    quote_word.extend(word_list)

# ����ͼ
# img_path = r'D:\Pycharm\Project\Ploting\img'
#
# # img_save = os.path.join(img_path, '1.png')
#
# text = ' '.join(quote_word)
#
# test_mask = np.array(Image.open(r'D:\Pycharm\Project\Ploting\picture\dog.png'))
# wordcloud = WordCloud(font_path=r'C:\Windows\Fonts\simkai.ttf', collocations=False, width=1200, height=980,
#                       mask=test_mask, scale=15).generate(text)
#
# plt.imshow(wordcloud, interpolation="bilinear")
# plt.axis("off")
#
# # plt.savefig(img_save)
# plt.show()

# top20��Ƶͳ��
x = []
y = []
mycount = Counter(quote_word)
for key, val in mycount.most_common(20):  # ���򣨷���ǰ100����
    x.append(key)
    y.append(val)

fig, ax = plt.subplots()
ax.barh(x, y, color="red")
labels = ax.get_xticklabels()
plt.setp(labels, rotation=0, horizontalalignment='right')

for a, b in zip(x, y):
    plt.text(b+1, a, b, ha='center', va='center')
ax.legend(["label"], loc="lower right")

plt.rcParams['font.sans-serif'] = ['SimHei']  # ����������ʾ���ı�ǩ
plt.ylabel('����')
plt.xlabel('��Ƶ')
plt.rcParams['savefig.dpi'] = 300  # ͼƬ����
plt.rcParams['figure.dpi'] = 300  # �ֱ���
plt.rcParams['figure.figsize'] = (15.0, 8.0)  # �ߴ�
plt.title("���ı���������100��Ƶͳ��")
plt.show()
# top20��Ƶͳ��